"""
LECTOR DE PLANILLAS DE PAGO
Extrae datos de PDFs de planillas usando OCR local (sin Google)
"""

import streamlit as st
import pandas as pd
import io
import json
import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Importar validador de base local
try:
    from utils.validador_base_local import cargar_bases_locales, buscar_en_base
except ImportError:
    cargar_bases_locales = None
    buscar_en_base = None


# Caché para las bases locales
@st.cache_resource
def obtener_bases_locales():
    """Carga las bases locales una sola vez"""
    if cargar_bases_locales:
        return cargar_bases_locales()
    return {}


def generar_excel_local(df):
    """Genera archivo Excel localmente (función respaldo)"""
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Datos"
        
        # Título principal
        ws.merge_cells('A1:I1')
        titulo = ws['A1']
        titulo.value = "PLANTILLA PAGOS REDIRECCIONAMIENTO"
        titulo.font = Font(bold=True, size=14, color="FFFFFF")
        titulo.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        titulo.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 25
        
        # Reordenar columnas: RAZON_SOCIAL primero
        columnas_orden = ['Archivo', 'RAZON_SOCIAL']
        columnas_restantes = [col for col in df.columns if col not in columnas_orden]
        df_ordenado = df[columnas_orden + columnas_restantes] if 'RAZON_SOCIAL' in df.columns else df
        
        # Encabezados (fila 2)
        for col_idx, column_title in enumerate(df_ordenado.columns, 1):
            cell = ws.cell(row=2, column=col_idx)
            cell.value = column_title
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Datos (a partir de fila 3)
        for r_idx, row in enumerate(df_ordenado.values, 3):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.value = value
                cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # Ajustar ancho de columnas (evitar celdas mergeadas)
        for col_idx in range(1, len(df_ordenado.columns) + 1):
            max_length = 0
            column_letter = chr(64 + col_idx) if col_idx < 27 else 'A' + chr(64 + col_idx - 26)
            
            for row_idx in range(2, len(df_ordenado) + 3):
                try:
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if cell.value and not isinstance(cell, type(None)):
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
    except Exception as e:
        return None


def extraer_texto_pdf(pdf_content):
    """Extrae texto del PDF usando PyPDF2 o pdfplumber"""
    try:
        # Intentar 1: PyPDF2 (extrae texto directo si está incrustado)
        try:
            from PyPDF2 import PdfReader
            
            pdf_file = io.BytesIO(pdf_content)
            reader = PdfReader(pdf_file)
            
            texto_completo = ""
            for pagina in reader.pages[:10]:  # Primeras 10 páginas
                texto = pagina.extract_text()
                if texto:
                    texto_completo += texto + "\n"
            
            if texto_completo and len(texto_completo.strip()) > 50:
                return texto_completo
        except Exception as e_pypdf:
            pass
        
        # Intentar 2: pdfplumber (alternativa robusta)
        try:
            import pdfplumber
            
            pdf_file = io.BytesIO(pdf_content)
            
            with pdfplumber.open(pdf_file) as pdf:
                texto_completo = ""
                for pagina in pdf.pages[:10]:  # Primeras 10 páginas
                    texto = pagina.extract_text()
                    if texto:
                        texto_completo += texto + "\n"
                
                if texto_completo and len(texto_completo.strip()) > 50:
                    return texto_completo
        except Exception as e_pdfplumber:
            pass
        
        return ""
    
    except Exception as e:
        st.error(f"Error extrayendo texto: {e}")
        return ""


def limpiar_periodo(periodo_str):
    """Elimina guiones del período: '2012-12' -> '201212'"""
    if periodo_str == "No detectado" or not periodo_str:
        return "No detectado"
    return periodo_str.replace('-', '').replace(' ', '')


def extraer_campo(texto, patron):
    """Extrae un campo del texto usando regex"""
    try:
        match = re.search(patron, texto, re.IGNORECASE | re.MULTILINE | re.DOTALL)
        if match:
            valor = match.group(1).strip()
            return valor if valor else "No detectado"
        return "No detectado"
    except:
        return "No detectado"


def extraer_afiliados(texto):
    """
    Extrae TODOS los afiliados de la tabla en el PDF
    Retorna lista de diccionarios con datos de cada afiliado
    """
    try:
        # Patrón para encontrar filas de afiliados en la tabla
        # Estructura: Nro | CUSPP | Nombre | Remuneración | ...
        patron = r'^\s*(\d+)\s+([0-9]{6}[A-Z]{5}\d)\s+([A-Z\s,\.]+?)(?=\s+[SN]\s)'
        
        afiliados = []
        lineas = texto.split('\n')
        
        for i, linea in enumerate(lineas):
            match = re.match(patron, linea.strip())
            if match:
                afiliados.append({
                    'nro': match.group(1).strip(),
                    'cussp': match.group(2).strip(),
                    'nombre': match.group(3).strip()
                })
        
        return afiliados if afiliados else []
    except:
        return []


def calcular_monto_total_planilla(texto):
    """
    Calcula el MONTO TOTAL de la planilla (suma de todos los afiliados)
    Retorna string formateado en soles
    """
    try:
        monto_fondo = extraer_campo(texto, r'Total\s+Fondo\s+Pensiones[\s\n]+S/\.[\s\n]+([\d.]+)')
        
        matches = re.findall(r'Retenciones(?:\s+y)?\s+Retribuciones[\s\n]+S/\.[\s\n]+([\d.]+)', texto, re.IGNORECASE | re.MULTILINE | re.DOTALL)
        monto_retenciones = matches[-1] if matches else "No detectado"
        
        def limpiar_monto(monto_str):
            if monto_str == "No detectado" or not monto_str:
                return 0.0
            try:
                monto_limpio = monto_str.replace(' ', '').replace('\n', '').strip()
                return float(monto_limpio)
            except:
                return 0.0
        
        valor_fondo = limpiar_monto(monto_fondo)
        valor_retenciones = limpiar_monto(monto_retenciones)
        
        total = valor_fondo + valor_retenciones
        if total > 0:
            return f"S/. {total:.2f}"
        
        return "No detectado"
    except:
        return "No detectado"



# Configurar página
st.set_page_config(
    page_title="Lector de Planillas",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

st.title("📊 Lector de Planillas de Pago")
st.markdown("Extrae datos automáticamente de PDFs de planillas usando OCR local")

# Sidebar
st.sidebar.markdown("## ⚙️ Configuración")

datos_cliente = {}
datos_cliente["nombre"] = st.sidebar.text_input("Nombre de tu empresa:", "")
datos_cliente["ruc"] = st.sidebar.text_input("RUC:", "")

# Campos a extraer
st.sidebar.markdown("## 📋 Campos a extraer")
campos_a_mostrar = st.sidebar.multiselect(
    "Selecciona los campos que deseas ver:",
    [
        "RUC", "RAZON_SOCIAL", "PERIODO", "CUSSP", "AFILIADO",
        "FECHA_PAGO", "N_PLANILLA", "MONTO", "OBSERVACION"
    ],
    default=["RUC", "RAZON_SOCIAL", "PERIODO", "CUSSP", "AFILIADO",
             "FECHA_PAGO", "N_PLANILLA", "MONTO", "OBSERVACION"]
)

# Área principal
st.markdown("---")

# Cargar archivos
st.markdown("### 📁 Carga tus Planillas (PDF)")

archivos_cargados = st.file_uploader(
    "Selecciona PDF(s) de planillas",
    type=["pdf"],
    accept_multiple_files=True,
    help="Puedes cargar múltiples archivos PDF"
)

if archivos_cargados:
    st.markdown("---")
    st.markdown("### 🔄 Procesando...")
    
    datos_extraidos = []
    progress_bar = st.progress(0)
    status_text = st.empty()
    
    for idx, archivo in enumerate(archivos_cargados):
        try:
            # Actualizar progreso
            progreso = (idx + 1) / len(archivos_cargados)
            progress_bar.progress(progreso)
            status_text.text(f"Procesando {idx+1}/{len(archivos_cargados)}: {archivo.name}")
            
            # Leer contenido del PDF
            pdf_content = archivo.read()
            
            # Intentar extraer texto
            texto = extraer_texto_pdf(pdf_content)
            
            if texto and len(texto.strip()) > 50:
                st.success(f"✅ Texto extraído de {archivo.name}")
                
                # Extraer datos generales de la planilla
                ruc_val = extraer_campo(texto, r'RUC[:\s]+(\d{11})')
                periodo_val = limpiar_periodo(extraer_campo(texto, r'Periodo\s+(?:de\s+Devengue)?[:\s]+(\d{4}-\d{2})'))
                
                monto_total = calcular_monto_total_planilla(texto)
                
                # DEBUG: Mostrar valores extraídos
                with st.expander(f"🔍 Debug - {archivo.name}", expanded=False):
                    st.write(f"**RUC extraído:** {ruc_val}")
                    st.write(f"**PERÍODO (limpio):** {periodo_val}")
                    st.write(f"**MONTO:** {monto_total}")
                    
                    # Extraer AFILIADOS para debug
                    afiliados_debug = extraer_afiliados(texto)
                    st.write(f"**Afiliados encontrados en PDF:** {len(afiliados_debug)}")
                    if afiliados_debug:
                        for aff in afiliados_debug:
                            st.write(f"  - {aff['nombre']} ({aff['cussp']})")
                    else:
                        st.write("❌ No se encontraron afiliados en tabla")
                
                datos_base = {
                    "Archivo": archivo.name,
                    "RUC": ruc_val,
                    "RAZON_SOCIAL": extraer_campo(texto, r'(?:Nombre\s+o\s+)?Razón\s+Social[:\s]+([^\n]+?)(?:\s*RUC|$)'),
                    "PERIODO": periodo_val,
                    "FECHA_PAGO": extraer_campo(texto, r'Fecha\s+de\s+Pago[:\s]*\n?\s*(\d{1,2}[/-]\d{1,2}[/-]\d{4})'),
                    "N_PLANILLA": extraer_campo(texto, r'(?:Número\s+de\s+)?Planilla[:\s]+(\d+)'),
                    "MONTO": monto_total,  # MONTO TOTAL COMPARTIDO ENTRE AFILIADOS
                    "OBSERVACION": ""
                }
                
                # Cargar bases locales para validación (solo como fallback)
                bases_locales = obtener_bases_locales() if buscar_en_base else None
                
                # Extraer TODOS los afiliados del PDF
                afiliados = extraer_afiliados(texto)
                
                if afiliados:
                    # Caso 1: Se encontraron afiliados en el PDF
                    # Crear una fila por cada afiliado con MONTO compartido (sin validación)
                    st.info(f"✅ Se encontraron {len(afiliados)} afiliado(s) en {archivo.name}")
                    
                    for idx, afiliado in enumerate(afiliados):
                        fila = datos_base.copy()
                        fila["CUSSP"] = afiliado['cussp']
                        fila["AFILIADO"] = afiliado['nombre']
                        # OBSERVACION vacío porque se encontraron datos en el PDF
                        fila["OBSERVACION"] = ""
                        # MONTO solo en la primera fila
                        if idx > 0:
                            fila["MONTO"] = ""
                        datos_extraidos.append(fila)
                else:
                    # Caso 2: No se encontró tabla de afiliados
                    # Intentar extraer CUSSP y AFILIADO del PDF
                    cussp_value = extraer_campo(texto, r'^\s*1\s+([0-9]{6}[A-Z]{5}\d)') or \
                                  extraer_campo(texto, r'CUSPP[\s\S]*?(\d{6}[A-Z]{5}\d)')
                    afiliado_name = extraer_campo(texto, r'[0-9]{6}[A-Z]{5}\d\s+([A-Z\s,\.]+?)(?=\s+S\s|\s+N\s)')
                    
                    if cussp_value != "No detectado" and afiliado_name != "No detectado":
                        # Caso 2a: CUSSP y AFILIADO encontrados en el PDF
                        datos_base["CUSSP"] = cussp_value
                        datos_base["AFILIADO"] = afiliado_name
                        datos_base["OBSERVACION"] = ""
                        datos_extraidos.append(datos_base)
                        st.info(f"✅ Se extrajo 1 afiliado de {archivo.name}")
                    else:
                        # Caso 2b: NO se encontraron CUSSP y AFILIADO en el PDF
                        # FALLBACK: Buscar en la base local
                        if bases_locales:
                            try:
                                # Buscar todos los registros que coincidan con DOCUMENTO y PERIODO
                                validacion = buscar_en_base(
                                    ruc=ruc_val,
                                    documento=ruc_val,
                                    periodo=periodo_val,
                                    cussp="",
                                    afiliado_pdf="",
                                    bases=bases_locales
                                )
                                
                                if validacion['encontrado'] and validacion.get('afiliados'):
                                    # Se encontraron múltiples afiliados en la base local
                                    # Crear una fila por cada afiliado con MONTO compartido
                                    for idx, aff in enumerate(validacion['afiliados']):
                                        fila = datos_base.copy()
                                        fila["CUSSP"] = aff['cussp']
                                        fila["AFILIADO"] = aff['afiliado']
                                        fila["OBSERVACION"] = aff['observacion']
                                        # MONTO solo en la primera fila
                                        if idx > 0:
                                            fila["MONTO"] = ""
                                        datos_extraidos.append(fila)
                                    st.info(f"✅ Se obtuvieron {len(validacion['afiliados'])} afiliado(s) de la base local para {archivo.name}")
                                else:
                                    # No se encontró en bases
                                    datos_base["CUSSP"] = "No detectado"
                                    datos_base["AFILIADO"] = "No detectado"
                                    datos_base["OBSERVACION"] = "No se encontró en PDF ni en bases locales"
                                    datos_extraidos.append(datos_base)
                                    st.warning(f"⚠️ No se pudieron extraer datos de {archivo.name}")
                            except Exception as e:
                                datos_base["CUSSP"] = "No detectado"
                                datos_base["AFILIADO"] = "No detectado"
                                datos_base["OBSERVACION"] = f"Error consultando bases: {str(e)}"
                                datos_extraidos.append(datos_base)
                                st.warning(f"⚠️ Error al buscar en bases para {archivo.name}")
                        else:
                            # Sin bases locales disponibles
                            datos_base["CUSSP"] = "No detectado"
                            datos_base["AFILIADO"] = "No detectado"
                            datos_base["OBSERVACION"] = "No se encontró en PDF y bases locales no disponibles"
                            datos_extraidos.append(datos_base)
                            st.warning(f"⚠️ No se encontraron datos en {archivo.name}")
            else:
                st.warning(f"⚠️ No se extrajo texto de {archivo.name}")
        
        except Exception as e:
            st.error(f"❌ Error procesando {archivo.name}: {str(e)}")
    
    # Mostrar resultados
    if datos_extraidos:
        st.markdown("---")
        st.success(f"✅ Se procesaron {len(datos_extraidos)} fila(s) en total")
        
        # DataFrame
        df = pd.DataFrame(datos_extraidos)
        
        # Mostrar tabla
        st.markdown("### 📋 Datos Extraídos")
        if campos_a_mostrar:
            cols_mostrar = ["Archivo"] + [c for c in campos_a_mostrar if c in df.columns]
            st.dataframe(df[cols_mostrar], use_container_width=True)
        else:
            st.dataframe(df, use_container_width=True)
        
        # Descargar Excel
        try:
            from utils.excel_generator import generar_excel
        except ImportError:
            generar_excel = generar_excel_local
        
        col1, col2 = st.columns(2)
        
        with col2:
            excel_file = generar_excel(df)
            if excel_file:
                st.download_button(
                    label="📥 Descargar Excel",
                    data=excel_file,
                    file_name=f"PLANTILLA_PAGOS_REDIRECCIONAMIENTO_{datos_cliente['nombre'].replace(' ', '_') if datos_cliente['nombre'] else 'exportacion'}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
    else:
        st.error("❌ No se pudo extraer datos de los archivos")
else:
    st.info("👆 Carga archivos PDF para comenzar")

# Footer
st.markdown("---")
st.markdown("""
<div style='text-align: center; color: #888; font-size: 12px;'>
    Lector de Planillas © 2025 | Extractor con OCR local
</div>
""", unsafe_allow_html=True)
