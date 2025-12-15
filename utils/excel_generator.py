"""
Generador de archivos Excel
"""
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


def generar_excel(df):
    """
    Genera archivo Excel a partir de DataFrame
    
    Args:
        df: DataFrame con los datos
    
    Returns:
        bytes: Archivo Excel
    """
    try:
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Datos"
        
        # Reordenar columnas según el orden deseado
        orden_columnas = [
            'RUC', 'RAZON_SOCIAL', 'PERIODO', 'CUSSP', 'AFILIADO',
            'FECHA_PAGO', 'N_PLANILLA', 'MONTO', 'OBSERVACION'
        ]
        # Mantener solo las columnas que existan en el DataFrame
        columnas_existentes = [col for col in orden_columnas if col in df.columns]
        # Agregar columnas restantes (como Archivo)
        columnas_restantes = [col for col in df.columns if col not in columnas_existentes]
        df_ordenado = df[columnas_existentes + columnas_restantes]
        
        # Calcular el rango de columnas para el merge dinámicamente
        num_cols = len(df_ordenado.columns)
        col_letra_fin = chr(64 + num_cols) if num_cols < 27 else 'A' + chr(64 + num_cols - 26)
        rango_merge = f'A1:{col_letra_fin}1'
        
        # Título principal
        ws.merge_cells(rango_merge)
        titulo = ws['A1']
        titulo.value = "PLANTILLA PAGOS REDIRECCIONAMIENTO"
        titulo.font = Font(bold=True, size=14, color="FFFFFF")
        titulo.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        titulo.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 25
        
        # Encabezados (fila 2)
        for col_idx, column_title in enumerate(df_ordenado.columns, 1):
            cell = ws.cell(row=2, column=col_idx)
            cell.value = column_title
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Datos (a partir de fila 3)
        # Encontrar índice de columna MONTO para aplicar formato de moneda
        monto_col_idx = None
        for col_idx, col_name in enumerate(df_ordenado.columns, 1):
            if col_name == 'MONTO':
                monto_col_idx = col_idx
                break
        
        for r_idx, row in enumerate(df_ordenado.values, 3):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.value = value
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
                # Aplicar formato de moneda a la columna MONTO
                if c_idx == monto_col_idx and isinstance(value, (int, float)) and value > 0:
                    cell.number_format = '_("S/. "* #,##0.00_);_("S/. "* (#,##0.00);_("S/. "* "-"??_);_(@_)'
        
        # Ajustar ancho de columnas (evitar celdas mergeadas)
        for col_idx in range(1, len(df_ordenado.columns) + 1):
            max_length = 0
            column_letter = chr(64 + col_idx) if col_idx < 27 else 'A' + chr(64 + col_idx - 26)
            
            for row_idx in range(2, len(df_ordenado) + 3):
                try:
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if cell.value and not isinstance(cell, type(None)):
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Guardar en bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
    
    except Exception as e:
        raise Exception(f"Error al generar Excel: {str(e)}")
